import sys

import openpyxl
import openpyxl.utils

ICARE_DATA_PARSE_START = 4

def iCare_parse_columns(file_name):
    book = openpyxl.load_workbook(filename = file_name, read_only = True)

    sheet = book.active

    columns = []
    row = 3

    for i in range(1, sheet.max_column + 1):
        cell = "{}{}".format(openpyxl.utils.get_column_letter(i), row)
        columns.append(sheet[cell].value)

    return columns

def parse_xlsx(file_name, column_names):
    '''(str, list of str, int, int) -> list of list of str
    
    '''
    column_len = len(column_names)
    if (column_len <= 0):
        return []

    book = openpyxl.load_workbook(filename = file_name, read_only = True)
    sheet = book.active

    rows = aggregate_xlsx_rows(sheet)

    return rows

def aggregate_xlsx_rows(sheet):
    values = []
    for row in range(ICARE_DATA_PARSE_START, sheet.max_row + 1):
        values.append([cell.value for cell in sheet[row]])
    return values
